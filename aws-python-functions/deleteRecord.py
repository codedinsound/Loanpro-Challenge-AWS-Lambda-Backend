import boto3
import io
import abc
import json 
from openpyxl import load_workbook

# ===============================================================
# Database Logic For Python 
# ===============================================================
list_of_session_tokens = {'abc1', 'abc2', 'abc3', 'abc4', 'abc5'}

class Database(abc.ABC):
    @abc.abstractclassmethod
    def connect(self):
        pass
    @abc.abstractclassmethod
    def disconnect(self):
        pass
    @abc.abstractclassmethod
    def query(self, q = ''):
        pass

class DatabaseManagementController:
    def __init__(self):
        self.database = None

    def connect(self):
        self.database.connect()
    
    def disconnect(self):
        del self.database
    
    def query(self, q = ''): 
        return self.database.query(q)

    def setDatabase(self, database = Database):
        self.database = database
        return self

    @staticmethod
    def instance():
        return DatabaseManagementController()

class ExcelSheetTestDatabase(Database):
    def __init__(self):
        self.s3 = boto3.client('s3')
        self.workbook = None 
        self.sheet_names = []
        self.current_sheet = None

    def __del__(self):
        print("destroying object....")

    # Private Methods 
    def __search(self, record_id = ''):
        print(record_id)
        record_index = 0
        res = {'status': 'failed', 'response': {}}
        for row in self.current_sheet.rows:
            if (row[0].value == record_id):
                res['status'] = 'success'
                res['response'] = {
                    'record_index': record_index,
                    'record': row
                }
                return res
            record_index += 1
        return res 

    def __set_worksheet(self, sheet_selected = ''): 
        res = {'status': 'success', 'response': 'selected sheet set'}
        if (sheet_selected in self.sheet_names):
            self.current_sheet = self.workbook[sheet_selected]
        else:
            res = {'status': 'failed', 'response': 'could not set sheet'}
        return res 

    def __update_workbook(self, update=''):
        print("updating workbook") 
        file_to_upload = io.BytesIO()
        self.workbook.save(file_to_upload)
        file_to_upload.seek(0)
        self.s3.upload_fileobj(file_to_upload, "loanpro-challenge-aws-la-serverlessdeploymentbuck-1mo1bm3wp9e2s", "test-data/test3.xlsx")

    def __get_user_transaction_records(self, user_id = ''):
        print(user_id)
        id = int(user_id)
        res = {'status': 'failed', 'response': None}
        user_records = []
        for row in self.current_sheet.iter_rows(max_col=8, values_only=True):
            if (row[2] == id and row[7] != 'T'):
                user_records.append({
                    'id': row[0],
                    'operation_id': row[1],
                    'date': row[6],
                    'amount': row[3],
                    'user_balance': row[4]
                })
        res['status'] = 'success'
        res['response'] = user_records
        return res 
            
    # Public Methods 
    def connect(self):
        response = self.s3.get_object(Bucket="loanpro-challenge-aws-la-serverlessdeploymentbuck-1mo1bm3wp9e2s", Key="test-data/users_db.xlsx")
        file_stream = response['Body']
        self.workbook = load_workbook(filename = io.BytesIO(file_stream.read()))
        self.sheet_names = self.workbook.sheetnames

    def disconnect(self):
        print("disconnecting from db....")
        
    def query(self, q = ''):
        db_res = {'status': '', 'response': {}}
        queries = q.split()
        db_function_calls = {
            'GET': self.__get_user_transaction_records,
            'SEARCH': self.__search,
            'SET_WORKSHEET': self.__set_worksheet,
            'UPDATE_WB': self.__update_workbook
        }
        if (len(queries) > 0 and queries[0] in db_function_calls):
            db_res = db_function_calls.get(queries[0])(queries[1])
        else:
            db_res = {'status': 'failed', 'response': 'invalid query...'}
        return db_res 

# MARK: Check if client session token is valid
def confirm_client_session_is_alive(userID, sessionToken): 
    print(userID, sessionToken)
    is_valid_session = False
    if (sessionToken in list_of_session_tokens): is_valid_session = True 
    return is_valid_session

# Mark: Check if client POST reponse is valid
def confirm_valid_client_reponse(clientResponse):
    is_valid_response = False 
    if ('operation' in clientResponse and 'sessionToken' in clientResponse and 'userID' in clientResponse and 'record' in clientResponse and clientResponse['operation'] == 'DELETE_RECORD'):
        is_valid_response = confirm_client_session_is_alive(clientResponse['userID'], clientResponse['sessionToken'])
    return is_valid_response

# MARK: Retrieve User Records from Workbook
def retrieve_user_transaction_records(clientResponse, db = Database):
    user_records = []
    db.query('SET_WORKSHEET UsersOperationsRecords')
    res = db.query('GET ' + str(clientResponse['userID']))
    if (res['status'] == 'success'):
        user_records = res['response']
    return user_records

# MARK: Handle client deletion of record 
def soft_delete_transaction_record(clientResponse, db = Database):
    print("Soft Deleting Record")
    db.query('SET_WORKSHEET UsersOperationsRecords')
    res = db.query('SEARCH ' + str(clientResponse['record']['id']))
    if (res['status'] == 'success'):
        res['response']['record'][7].value = 'T'
        db.query('UPDATE_WB _')

# MARK: AWS Lambda Handler 
def lambda_handler(event, context):
    proceed = False 
    proceed = confirm_valid_client_reponse(event)
    response = {}
    if (proceed): 
        db = DatabaseManagementController.instance().setDatabase(ExcelSheetTestDatabase())
        db.connect()
        soft_delete_transaction_record(event, db)
        user_records = retrieve_user_transaction_records(event, db)
        db.disconnect()
        response = {
            'statusCode': 200,
            'body': json.dumps(user_records)
        }
    else:
        response = {
            'statusCode': 418, 
            'body': {}
        }
    return response

# x = {'userID': 4, 'sessionToken': 'abc1', 'operation': 'DELETE_RECORD', 'record': {'id': '4-2', 'operation_id': 'ADD', 'date': '2023-02-12T03:57:31.986Z', 'amount': 5000
# , 'user_balance': 97925}}
# received_payload = lambda_handler(x, '')
# print(received_payload)
